import gspread
from oauth2client.service_account import ServiceAccountCredentials
import hashlib
from datetime import datetime
from zoneinfo import ZoneInfo
import json
import uuid
import time
import re
import unicodedata

# Printer
from escpos.printer import Network
from decimal import Decimal, ROUND_HALF_UP

BUSINESS_TZ = ZoneInfo("America/Guayaquil")


class ReceiptPrinter: 
    def __init__(self):
        # Configure for RPT004 - adjust vendor/product ID for your printer
        # To find IDs: lsusb (Linux) or Device Manager (Windows)
        try:
            # USB Printer
            #self.printer = Usb(0x04b8, 0x0e14)  # Replace with your RPT004 IDs
            
            # OR Network Printer (if using WiFi/Ethernet)
            self.printer = Network("192.168.1.100")
            
            # OR File Printer (for testing - prints to file)
            # self.printer = File("/dev/usb/lp0")
            
        except Exception as e:
            print(f"Printer initialization error: {e}")
            self.printer = None
    
    def print_receipt(self, receipt_data):
        """Print receipt to thermal printer"""
        if not self.printer:
            return {'success': False, 'error': 'Printer not initialized'}
        
        try:
            business = receipt_data['business']
            sale = receipt_data['sale']
            items = receipt_data['items']
            totals = receipt_data['totals']
            
            # Set encoding
            self.printer.charcode('USA')
            
            # Header - Centered
            self.printer.set(align='center', text_type='B', width=2, height=2)
            self.printer.text(f"{business['name']}\n")
            
            self.printer.set(align='center', text_type='normal', width=1, height=1)
            self.printer.text(f"{business['address']}\n")
            self.printer.text(f"{business['RUC']}\n")

            # Separator
            self.printer.text("================================\n")
            
            # Sale Info - Left aligned
            self.printer.set(align='left')
            self.printer.text(f"Fecha: {sale['fecha']} {sale['hora']}\n")          
            self.printer.text("--------------------------------\n")
            
            # Items Header
            self.printer.set(text_type='B')
            self.printer.text(f"{'Producto':<20} {'Cant':>4} {'Total':>8}\n")
            self.printer.set(text_type='normal')
            self.printer.text("--------------------------------\n")
            
            # Items
            for item in items:
                # Product name (can wrap if long)
                name = item['product_name'][:20]
                self.printer.text(f"{name:<20}\n")
                
                # Quantity, price, total
                qty = item['quantity_sold']
                price = item['price']
                total = price * qty
                self.printer.text(f"  ${price:.2f} x {qty:>2}        ${total:>7.2f}\n")
            
            self.printer.text("================================\n")
            
            # Totals
            self.printer.set(text_type='B', width=2, height=2)
            self.printer.text(f"TOTAL:          ${totals['total']:>8.2f}\n")
            
            #self.printer.set(text_type='normal', width=1, height=1)
            #if totals['received'] > 0:
            #    self.printer.text(f"Recibido:       ${totals['received']:>8.2f}\n")
            #    self.printer.text(f"Cambio:         ${totals['change']:>8.2f}\n")
            
            self.printer.text("--------------------------------\n")
            
            # Footer - Centered
            self.printer.set(align='center')
            self.printer.text("\n")
            self.printer.set(text_type='B')
            self.printer.text("¡Gracias por su compra!\n")
            self.printer.text("\n")
            
            # Cut paper
            self.printer.cut()
            
            return {'success': True, 'message': 'Receipt printed successfully'}
            
        except Exception as e:
            return {'success': False, 'error': str(e)}


class InventoryManager:
    def __init__(self, credentials_file, spreadsheet_name):
        scope = ['https://spreadsheets.google.com/feeds',
                 'https://www.googleapis.com/auth/drive']
        
        creds = ServiceAccountCredentials.from_json_keyfile_name(
            credentials_file, scope
        )
        self.client = gspread.authorize(creds)
        self.spreadsheet = self.client.open(spreadsheet_name)
        self.sheet_inventory = self.spreadsheet.worksheet('Inventario')
        self.sheet_sales = self.spreadsheet.worksheet('Ventas')
        self.sheet_users = self.spreadsheet.worksheet('Usuarios')  # Nueva hoja
        self._row_cache = {}
        self._row_cache_timestamp = 0
        self._row_cache_ttl = 300

    def _get_product_row(self, product_code):
        """Get row number from cache or fetch from sheet"""
        now = time.time()
        
        # Refresh entire cache if expired
        if (now - self._row_cache_timestamp) > self._row_cache_ttl:
            print("Refreshing row cache...")
            codes = self.sheet_inventory.col_values(2)  # fetch only codigo column
            self._row_cache = {code: i + 1 for i, code in enumerate(codes) if code}
            self._row_cache_timestamp = now
            print(f"Cache loaded with {len(self._row_cache)} products")
        
        return self._row_cache.get(product_code)

    def hash_password(self, password):
        """Hash de contraseña con SHA256"""
        return hashlib.sha256(password.encode()).hexdigest()
    
    def create_user(self, username, password, role='vendedor', nombre=''):
        """Crea un nuevo usuario"""
        try:
            # Verificar si el usuario ya existe
            users = self.sheet_users.get_all_records()
            for user in users:
                if user['Usuario'].lower() == username.lower():
                    return {
                        'success': False,
                        'message': 'El usuario ya existe'
                    }
            
            # Hash de la contraseña
            hashed_password = self.hash_password(password)
            
            # Obtener siguiente ID
            next_id = len(users) + 1
            
            # Crear usuario
            row = [
                next_id,
                username,
                hashed_password,
                role,  # admin, vendedor, cajero
                nombre,
                'Si',  # Activo
                ''  # UltimoAcceso
            ]
            
            self.sheet_users.append_row(row)
            
            return {
                'success': True,
                'message': 'Usuario creado exitosamente',
                'user_id': next_id
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def authenticate_user(self, username, password):
        """Autentica un usuario"""
        try:
            users = self.sheet_users.get_all_records()
            hashed_password = self.hash_password(password)
            
            for user in users:
                if (user['Usuario'].lower() == username.lower() and 
                    user['Password'] == hashed_password and 
                    user['Activo'].lower() == 'si'):
                    
                    # Actualizar último acceso
                    user_row = users.index(user) + 2  # +2 porque fila 1 es header y index empieza en 0
                    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    self.sheet_users.update_cell(user_row, 7, now)  # Columna 7 es UltimoAcceso
                    
                    return {
                        'success': True,
                        'user': {
                            'id': user['ID'],
                            'username': user['Usuario'],
                            'role': user['Rol'],
                            'nombre': user['Nombre']
                        }
                    }
            
            return {
                'success': False,
                'message': 'Usuario o contraseña incorrectos'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def get_all_users(self):
        """Obtiene todos los usuarios (sin passwords)"""
        try:
            users = self.sheet_users.get_all_records()
            users_list = []
            
            for user in users:
                users_list.append({
                    'id': user['ID'],
                    'username': user['Usuario'],
                    'role': user['Rol'],
                    'nombre': user['Nombre'],
                    'activo': user['Activo'],
                    'ultimo_acceso': user.get('UltimoAcceso', '')
                })
            
            return {
                'success': True,
                'users': users_list
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }

    def get_inventory(self):
        """Obtiene todo el inventario"""
        records = self.sheet_inventory.get_all_records()
        return records
    
    def add_product(self, product_data):
        """Agrega un nuevo producto a la hoja de Inventario"""
        try:
            now = datetime.now(BUSINESS_TZ)
            ultima_actualizacion = now.strftime('%Y-%m-%d %H:%M:%S')
            
            # Obtener el último ID para generar el siguiente
            all_records = self.sheet_inventory.get_all_records()
            next_id = len(all_records) + 1  # El siguiente ID es el total de registros + 1
            new_row = len(all_records) + 2  # +2 for header row

            
            # Preparar fila para insertar
            # Estructura: ID, Codigo, Nombre, Cantidad, Unidad, Costo, Precio_1, Precio_2, Precio_3, MinStock, UltimaActualizacion, Categoria, Subcategoria
            row = [
                next_id,
                product_data['codigo'],
                product_data['nombre'],
                product_data['cantidad'],
                product_data['unidad'],
                product_data['costo'],
                product_data['precio_1'],
                product_data['precio_2'],
                product_data['precio_3'],
                product_data['minStock'],
                ultima_actualizacion, 
                product_data.get('categoria', ''),
                product_data.get('subcategoria', ''),
            ]
            
            print(f'Producto para insertar: {row}')
            
            # Insertar el producto
            self.sheet_inventory.append_row(row)

            # Update cache
            self._row_cache[product_data['codigo']] = new_row
            print(f"Cache actualizado: {product_data['codigo']} -> row {new_row}")

            
            return {
                'success': True,
                'message': 'Producto agregado exitosamente',
                'product_code': product_data['codigo'],
                'product_id': next_id
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'message': f'Error al agregar producto: {str(e)}'
            }
    
    def update_product(self, product_code, updates):
        """Actualiza los campos de un producto existente (sin tocar cantidad)"""
        try:
            row = self._get_product_row(product_code)
            if row is None:
                return {'success': False, 'error': f'Producto {product_code} no encontrado'}

            now = datetime.now(BUSINESS_TZ).strftime('%Y-%m-%d %H:%M:%S')
            # Inventario columns: ID(1) Codigo(2) Nombre(3) Cantidad(4) Unidad(5)
            #   Costo(6) Precio_1(7) Precio_2(8) Precio_3(9) MinStock(10) UltimaActualizacion(11)
            #   Categoria(12) Subcategoria(13)
            batch = []
            if 'nombre' in updates:
                batch.append({'range': f'C{row}', 'values': [[updates['nombre']]]})
            if 'costo' in updates:
                batch.append({'range': f'F{row}', 'values': [[updates['costo']]]})
            if 'precio_1' in updates:
                batch.append({'range': f'G{row}', 'values': [[updates['precio_1']]]})
            if 'precio_2' in updates:
                batch.append({'range': f'H{row}', 'values': [[updates['precio_2']]]})
            if 'precio_3' in updates:
                batch.append({'range': f'I{row}', 'values': [[updates['precio_3']]]})
            if 'minStock' in updates:
                batch.append({'range': f'J{row}', 'values': [[updates['minStock']]]})
            if 'categoria' in updates:
                batch.append({'range': f'L{row}', 'values': [[updates['categoria']]]})
            if 'subcategoria' in updates:
                batch.append({'range': f'M{row}', 'values': [[updates['subcategoria']]]})
            batch.append({'range': f'K{row}', 'values': [[now]]})

            if batch:
                self.sheet_inventory.batch_update(batch)

            # Invalidate row cache so next read is fresh
            self._row_cache_timestamp = 0

            return {'success': True, 'message': 'Producto actualizado correctamente'}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def adjust_stock(self, product_code, cantidad_ajuste, motivo=''):
        """Ajusta el stock de un producto (puede ser positivo o negativo)"""
        try:
            row = self._get_product_row(product_code)
            if row is None:
                return {'success': False, 'error': f'Producto {product_code} no encontrado'}

            row_values = self.sheet_inventory.row_values(row)
            current_qty = float(row_values[3])
            new_qty = round(current_qty + float(cantidad_ajuste), 2)

            if new_qty < 0:
                return {'success': False, 'error': f'El ajuste dejaría el stock en {new_qty}, no puede ser negativo'}

            now = datetime.now(BUSINESS_TZ).strftime('%Y-%m-%d %H:%M:%S')
            self.sheet_inventory.batch_update([
                {'range': f'D{row}', 'values': [[new_qty]]},
                {'range': f'K{row}', 'values': [[now]]}
            ])

            # Log adjustment to Ajustes sheet
            try:
                ajustes_sheet = self._get_or_create_sheet('Ajustes', ['Fecha', 'Codigo', 'Nombre', 'AjusteAnterior', 'Ajuste', 'NuevoStock', 'Motivo'])
                ajustes_sheet.append_row([
                    now, product_code, row_values[2],
                    current_qty, cantidad_ajuste, new_qty, motivo
                ])
            except Exception:
                pass  # log failure is non-critical

            self._row_cache_timestamp = 0
            return {'success': True, 'new_quantity': new_qty, 'previous_quantity': current_qty}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def _get_or_create_sheet(self, title, headers):
        """Gets or creates a worksheet with the given title and headers"""
        try:
            return self.spreadsheet.worksheet(title)
        except Exception:
            sheet = self.spreadsheet.add_worksheet(title=title, rows=1000, cols=len(headers))
            sheet.append_row(headers)
            return sheet

    def get_product_by_code(self, code):
        """Busca un producto por código"""
        try:
            cell = self.sheet_inventory.find(code)
            row = self.sheet_inventory.row_values(cell.row)
            return {
                'id': row[0],
                'codigo': row[1],
                'nombre': row[2],
                'cantidad': int(row[3]),
                'precio': float(row[4]),
                'minStock': int(row[5])
            }
        except:
            return None
    
    def update_stock(self, product_code, quantity_sold, price_type):
        """Actualiza el stock después de una venta"""
        try:
            print(f"Updating {product_code}")
            # Buscar el producto
            row = self._get_product_row(product_code)

            if row is None:
                return {'success': False, 'error': f'Producto {product_code} no encontrado'}

            row_values = self.sheet_inventory.row_values(row)
                        
            # Obtener datos del producto
            product_id = row_values[0]
            product_name = row_values[2]
            current_qty = float(row_values[3])
            unidad = row_values[4].lower()
            price_1 = float(row_values[6])
            price_2 = float(row_values[7]) if row_values[7] else 0.0
            price_3 = float(row_values[8]) if row_values[8] else 0.0
            min_stock = float(row_values[9])
            quantity_sold = float(quantity_sold)
            
            if unidad == "unidad" and not quantity_sold.is_integer():
                return {
                    'success': False,
                    'error': 'Este producto solo se puede vender en unidades enteras'
                }

            # Verificar si hay suficiente stock
            if current_qty < quantity_sold:
                return {
                    'success': False,
                    'error': 'Stock insuficiente'
                }
            
            # Calcular nueva cantidad
            new_qty = round(current_qty - quantity_sold, 2)
            timestamp = datetime.now(BUSINESS_TZ).strftime('%Y-%m-%d %H:%M:%S')

            self.sheet_inventory.batch_update([
                {'range': f'D{row}', 'values': [[new_qty]]},
                {'range': f'K{row}', 'values': [[timestamp]]}
            ])
            
            # Verificar si requiere alerta
            alert = new_qty <= min_stock

            # Verificar el tipo de precio
            if price_type == "precio_2":
                selected_price = price_2
            elif price_type == "precio_3":
                selected_price = price_3
            else:
                selected_price = price_1
            
            return {
                    'success': True,
                    'product_id': product_id,
                    'product_code': product_code,
                    'product_name': product_name,
                    'price': selected_price,
                    'quantity_sold':quantity_sold,
                    'new_quantity': new_qty,
                    'alert': alert
                }
                
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
        
    def save_sale(self, sale_id, sale_details, total, vendedor='Sistema', metodo_pago='efectivo', referencia='', descuento=0):
        """Guarda el detalle de la venta en la hoja de Ventas"""

        try:
            now = datetime.now(BUSINESS_TZ)
            fecha = now.strftime('%Y-%m-%d')
            hora = now.strftime('%H:%M:%S')

            rows = []
            for item in sale_details:
                row = [
                    sale_id,
                    fecha,
                    hora,
                    item['product_id'],
                    item['product_code'],
                    item['product_name'],
                    item['quantity_sold'],
                    item['price'],
                    round(item['price'] * item['quantity_sold'], 2), # Subtotal
                    round(total, 2),                                  # Total
                    vendedor,                                         # Vendedor
                    metodo_pago,                                      # MetodoPago
                    referencia,                                       # Referencia
                    descuento,                                        # Descuento
                ]
                rows.append(row)
            
            print(f'Venta para insertar: {rows}')

            # Insertar todas las filas de la venta
            self.sheet_sales.append_rows(rows, value_input_option='USER_ENTERED')

            print(f'Venta guardada')
            
            return {
                'success': True,
                'sale_id': sale_id,
                'items_saved': len(rows)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
        
    def process_sale(self, cart_items, vendedor='Sistema', metodo_pago='efectivo', referencia='', descuento_porcentaje=0):
        """Procesa una venta completa"""

        sale_id = f"VTA-{datetime.now(BUSINESS_TZ).strftime('%Y%m%d')}-{str(uuid.uuid4())[:8]}"

        results = []
        alerts = []
        total_sale = 0
        sale_details = []
        
        # Procesar cada producto        
        for item in cart_items:
            result = self.update_stock(
                item['codigo'],
                item['cantidad_vendida'],
                item['tipoPrecio']
            )
            
            if not result['success']:
                return{
                    'success': False,
                    'error': f"Error en {item['codigo']}: {result['error']}"
                }
            
            results.append(result)

            # Calcular total
            subtotal = result['price']*result['quantity_sold']

            total_sale += subtotal
            
            # Guardar detalles para el historial
            sale_details.append({
                'product_id': result['product_id'],
                'product_code': result['product_code'],
                'product_name': result['product_name'],
                'price': result['price'],
                'quantity_sold': result['quantity_sold']
            })
            
            # Verificar alertas
            if result.get('alert'):
                alerts.append({
                    'producto': result['product_name'],
                    'cantidad_restante': result['new_quantity']
                })

        print("Inventario actualizado")
        
        # Apply discount
        descuento_porcentaje = max(0, min(100, float(descuento_porcentaje or 0)))
        descuento_monto = round(total_sale * (descuento_porcentaje / 100), 2)
        total_con_descuento = round(total_sale - descuento_monto, 2)

        save_result = self.save_sale(sale_id, sale_details, total_con_descuento, vendedor, metodo_pago, referencia, descuento_monto)

        if not save_result['success']:
            return {
                'success': False,
                'error': f"Venta procesada pero no se guardó en historial: {save_result['error']}"
            }

        return {
            'success': True,
            'sale_id': sale_id,
            'total': total_con_descuento,
            'subtotal': total_sale,
            'descuento': descuento_monto,
            'items': len(results),
            'results': results,
            'alerts': alerts
        }
    
    def refund_sale(self, sale_id, items):
        """
        Process a refund for a sale.
        items: list of {codigo, cantidad} dicts.
        Restores stock for each item and writes negative rows to Ventas sheet.
        """
        try:
            # Fetch the original sale rows to get price info
            all_sales = self.sheet_sales.get_all_records()
            original_rows = [r for r in all_sales if str(r.get('VentaID', '')) == sale_id]

            if not original_rows:
                return {'success': False, 'error': f'Venta {sale_id} no encontrada'}

            # Build price lookup from original sale
            price_lookup = {str(r['Codigo']): float(r.get('PrecioUnitario', 0)) for r in original_rows}

            refund_id = f"DEV-{datetime.now(BUSINESS_TZ).strftime('%Y%m%d')}-{str(uuid.uuid4())[:8]}"
            now = datetime.now(BUSINESS_TZ)
            fecha = now.strftime('%Y-%m-%d')
            hora = now.strftime('%H:%M:%S')

            refund_rows = []
            refunded_items = []
            total_refund = 0

            for item in items:
                codigo = str(item.get('codigo', '')).strip()
                cantidad = float(item.get('cantidad', 0))

                if not codigo or cantidad <= 0:
                    continue

                # Restore stock
                row = self._get_product_row(codigo)
                if row is None:
                    return {'success': False, 'error': f'Producto {codigo} no encontrado en inventario'}

                row_values = self.sheet_inventory.row_values(row)
                current_qty = float(row_values[3])
                new_qty = round(current_qty + cantidad, 2)
                timestamp = datetime.now(BUSINESS_TZ).strftime('%Y-%m-%d %H:%M:%S')
                self.sheet_inventory.batch_update([
                    {'range': f'D{row}', 'values': [[new_qty]]},
                    {'range': f'K{row}', 'values': [[timestamp]]}
                ])

                precio = price_lookup.get(codigo, 0)
                subtotal = round(-precio * cantidad, 2)
                total_refund += subtotal

                refund_rows.append([
                    refund_id, fecha, hora,
                    row_values[0],  # product_id
                    codigo,
                    row_values[2],  # nombre
                    -cantidad,
                    precio,
                    subtotal,
                    round(total_refund, 2),
                    'DEVOLUCION',
                    'devolucion',
                    sale_id,       # referencia = original sale ID
                    0,             # descuento
                ])
                refunded_items.append({'codigo': codigo, 'cantidad': cantidad, 'precio': precio})

            if refund_rows:
                self.sheet_sales.append_rows(refund_rows, value_input_option='USER_ENTERED')

            self._row_cache_timestamp = 0  # Invalidate cache

            return {
                'success': True,
                'refund_id': refund_id,
                'items_refunded': len(refund_rows),
                'total_refund': abs(total_refund),
            }
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def get_sales_history(self, limit=None, date_from=None, date_to=None):
        """Obtiene el historial de ventas con filtros opcionales"""
        try:
            records = self.sheet_sales.get_all_records()
            
            # Filtrar por fecha si se especifica
            if date_from:
                records = [r for r in records if r['Fecha'] >= date_from]
            if date_to:
                records = [r for r in records if r['Fecha'] <= date_to]
            
            # Limitar cantidad de resultados
            if limit:
                records = records[-limit:]
            
            return records
        except Exception as e:
            print(f"Error obteniendo historial: {e}")
            return []
    
    def get_sales_summary(self, date=None):
        """Obtiene un resumen de ventas del día"""
        try:
            if date is None:
                date = datetime.now(BUSINESS_TZ).strftime('%Y-%m-%d')
            
            records = self.sheet_sales.get_all_records()
            daily_sales = [r for r in records if r['Fecha'] == date]
            
            if not daily_sales:
                return {
                    'date': date,
                    'total_sales': 0,
                    'total_amount': 0,
                    'items_sold': 0,
                    'unique_sales': 0
                }
            
            # Calcular estadísticas
            unique_sales = len(set(r['VentaID'] for r in daily_sales))
            total_items = sum(r['Cantidad'] for r in daily_sales)
            total_amount = sum(r['Subtotal'] for r in daily_sales)
            
            return {
                'date': date,
                'total_sales': unique_sales,
                'total_amount': total_amount,
                'items_sold': total_items,
                'sales': daily_sales
            }
            
        except Exception as e:
            return {
                'error': str(e)
            }
            
    def _get_or_create_categories_sheet(self):
        try:
            return self.spreadsheet.worksheet('Categorias')
        except gspread.exceptions.WorksheetNotFound:
            sheet = self.spreadsheet.add_worksheet(title='Categorias', rows=200, cols=2)
            sheet.append_row(['Categoria', 'Subcategoria'])
            return sheet

    def get_categories(self):
        """Obtiene categorías y subcategorías desde la hoja Categorias"""
        try:
            sheet = self._get_or_create_categories_sheet()
            records = sheet.get_all_records()
            grouped = {}
            for row in records:
                cat = str(row.get('Categoria', '')).strip()
                sub = str(row.get('Subcategoria', '')).strip()
                if not cat:
                    continue
                if cat not in grouped:
                    grouped[cat] = []
                if sub and sub not in grouped[cat]:
                    grouped[cat].append(sub)
            return {'success': True, 'data': grouped}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def add_category(self, categoria, subcategoria=''):
        """Agrega una categoría o subcategoría a la hoja Categorias"""
        try:
            categoria = categoria.strip()
            subcategoria = subcategoria.strip() if subcategoria else ''
            if not categoria:
                return {'success': False, 'error': 'El nombre de la categoría no puede estar vacío'}
            sheet = self._get_or_create_categories_sheet()
            records = sheet.get_all_records()
            # Prevent exact duplicates
            for row in records:
                if (str(row.get('Categoria', '')).strip() == categoria and
                        str(row.get('Subcategoria', '')).strip() == subcategoria):
                    return {'success': False, 'error': 'Esta categoría/subcategoría ya existe'}
            sheet.append_row([categoria, subcategoria])
            return {'success': True, 'message': 'Categoría agregada exitosamente'}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def save_invoice_to_sheet(self, sale_id, invoice_result, cliente, cart_items, total):
        """Guarda los datos de una factura electrónica en la hoja 'Facturas'."""
        try:
            headers = [
                'FacturaID', 'VentaID', 'Fecha', 'Hora',
                'ClaveAcceso', 'NumeroFactura', 'NumAutorizacion', 'FechaAutorizacion',
                'ClienteID', 'ClienteNombre', 'ClienteEmail',
                'Total', 'Ambiente',
            ]
            sheet = self._get_or_create_sheet('Facturas', headers)
            now = datetime.now(BUSINESS_TZ)
            row = [
                f"FAC-{now.strftime('%Y%m%d%H%M%S')}",
                sale_id,
                now.strftime('%Y-%m-%d'),
                now.strftime('%H:%M:%S'),
                invoice_result.get('clave_acceso', ''),
                invoice_result.get('numero_factura', ''),
                invoice_result.get('numero_autorizacion', ''),
                invoice_result.get('fecha_autorizacion', ''),
                cliente.get('identificacion', ''),
                cliente.get('razon_social', ''),
                cliente.get('email', ''),
                total,
                invoice_result.get('ambiente', ''),
            ]
            sheet.append_row(row, value_input_option='USER_ENTERED')
            return {'success': True}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def send_invoice_email(self, cliente_email, cliente_nombre, invoice_result, xml_path=None):
        """Envía la factura autorizada al cliente por correo electrónico."""
        import smtplib
        import os
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders

        smtp_host = os.environ.get('SMTP_HOST', '')
        smtp_port = int(os.environ.get('SMTP_PORT', 587))
        smtp_user = os.environ.get('SMTP_USER', '')
        smtp_pass = os.environ.get('SMTP_PASS', '')

        if not smtp_host or not smtp_user or not smtp_pass:
            return {'success': False, 'error': 'Credenciales SMTP no configuradas (SMTP_HOST, SMTP_USER, SMTP_PASS)'}

        try:
            msg = MIMEMultipart()
            msg['From'] = smtp_user
            msg['To'] = cliente_email
            msg['Subject'] = f"Factura Electrónica – {invoice_result.get('numero_factura', '')}"

            body = (
                f"Estimado/a {cliente_nombre},\n\n"
                f"Adjuntamos su factura electrónica autorizada por el SRI.\n\n"
                f"Número: {invoice_result.get('numero_factura', '')}\n"
                f"Clave de acceso: {invoice_result.get('clave_acceso', '')}\n"
                f"Autorización: {invoice_result.get('numero_autorizacion', '')}\n"
                f"Fecha: {invoice_result.get('fecha_autorizacion', '')}\n\n"
                "Saludos,\nCentro Comercial TB"
            )
            msg.attach(MIMEText(body, 'plain', 'utf-8'))

            if xml_path and os.path.isfile(xml_path):
                with open(xml_path, 'rb') as f:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(f.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(xml_path)}"')
                    msg.attach(part)

            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, cliente_email, msg.as_string())

            return {'success': True}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def get_low_stock_alerts(self):
        """Obtiene todos los productos con stock bajo"""
        records = self.get_inventory()
        alerts = []
        
        for record in records:
            if record['Cantidad'] <= record['MinStock']:
                alerts.append({
                    'codigo': record['Codigo'],
                    'nombre': record['Nombre'],
                    'cantidad': record['Cantidad'],
                    'minimo': record['MinStock']
                })
        
        return alerts

    def get_profit_analysis(self, period='today', custom_start=None, custom_end=None):
        """Analiza las utilidades para cierre de caja por período"""
        try:
            from datetime import datetime, timedelta
            
            # Obtener ventas e inventario
            #sales = self.sheet_sales.get_all_records()
            #inventory = self.sheet_inventory.get_all_records()
            
            # Diccionario de costos
            #costs_dict = {item['Codigo']: float(item.get('Costo', 0)) for item in inventory}
            
            # Determinar rango de fechas según período
            now = datetime.now(BUSINESS_TZ)
            
            if period == 'today':
                print("Filtrando ventas de hoy")
                start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
                end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
                period_label = f"Hoy - {now.strftime('%d/%m/%Y')}"
                
            elif period == 'week':
                # Inicio de semana (lunes)
                print("Filtrando ventas de la semana")
                start_date = now - timedelta(days=now.weekday())
                start_date = start_date.replace(hour=0, minute=0, second=0)
                end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
                period_label = f"Esta Semana ({start_date.strftime('%d/%m')} - {end_date.strftime('%d/%m/%Y')})"
                
            elif period == 'month':
                # Inicio de mes
                print("Filtrando ventas del mes")
                start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
                period_label = f"Este Mes - {now.strftime('%B %Y')}"
                
            elif period == 'custom' and custom_start and custom_end:
                print("Filtrando ventas en rango personalizado")
                start_date = datetime.strptime(custom_start, '%Y-%m-%d').replace(tzinfo=BUSINESS_TZ)
                end_date = datetime.strptime(custom_end, '%Y-%m-%d').replace(hour=23, minute=59, second=59, tzinfo=BUSINESS_TZ)
                end_date = end_date.replace(hour=23, minute=59, second=59)
                period_label = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
            else:
                return {'success': False, 'error': 'Período no válido'}
            

            # Fetch only date column to find matching rows
            date_col = self.sheet_sales.col_values(2)[1:]  # skip header
            headers = self.sheet_sales.row_values(1)

            # Find matching row indices
            matching_rows = []
            for i, date_str in enumerate(date_col):
                try:
                    sale_date = datetime.strptime(date_str, '%Y-%m-%d').replace(tzinfo=BUSINESS_TZ)
                    if start_date.date() <= sale_date.date() <= end_date.date():
                        matching_rows.append(i + 2)  # +2: 1-indexed + skip header
                except:
                    continue

            if not matching_rows:
                return {
                    'success': True,
                    'data': {
                        'periodo': period_label,
                        'total_ingresos': 0,
                        'total_costos': 0,
                        'utilidad_neta': 0,
                        'margen_total': 0,
                        'total_ventas': 0,
                        'total_unidades': 0,
                        'ticket_promedio': 0,
                        'productos_vendidos': [],
                        'vendedores': [],
                        'ventas_detalle': [],
                        'metodo_pago_breakdown': [],
                    }
                }
            
            #print(f"Headers: {headers}")
            #print(f"First 5 dates in column: {date_col[:5]}")
            #print(f"Total rows in date col: {len(date_col)}")
            #print(f"Looking for dates between {start_date.date()} and {end_date.date()}")
            #print(f"Matching rows found: {matching_rows}")

            # Fetch only matching rows in one batch call
            col_letter = chr(64 + len(headers))
            sheet_name = self.sheet_sales.title  # 'Ventas'
            ranges = [f"'{sheet_name}'!A{row}:{col_letter}{row}" for row in matching_rows]            
            
            #print(f"Col letter: {col_letter}")
            #print(f"First range: {ranges[0]}")
            #print(f"Total ranges: {len(ranges)}")

            batch_data = self.sheet_sales.spreadsheet.values_batch_get(ranges)

            #print(f"Batch response keys: {batch_data.keys()}")
            #print(f"ValueRanges count: {len(batch_data.get('valueRanges', []))}")
            #print(f"First valueRange: {batch_data.get('valueRanges', [])[0] if batch_data.get('valueRanges') else 'EMPTY'}")
                                
            # Build records from batch response
            filtered_sales = []
            for value_range in batch_data.get('valueRanges', []):
                values = value_range.get('values', [])
                if values:
                    record = dict(zip(headers, values[0]))
                    filtered_sales.append(record)

            # Fetch inventory for costs
            inventory = self.sheet_inventory.get_all_records()
            costs_dict = {item['Codigo']: float(item.get('Costo', 0)) for item in inventory}

            # Calcular totales
            total_ingresos = Decimal("0.000")
            total_costos = Decimal("0.000")
            total_unidades = Decimal("0")
            ventas_detalle = []
            productos_vendidos = {}
            vendedores_stats = {}
            
            metodo_pago_stats = {}

            for sale in filtered_sales:
                codigo = sale.get('Codigo', '')
                cantidad = Decimal(str(sale.get('Cantidad', 0)))
                precio_venta = Decimal(str(sale.get('PrecioUnitario', 0)))
                costo_unitario = Decimal(str(costs_dict.get(codigo, 0)))
                vendedor = sale.get('Vendedor', 'Sistema')
                metodo_pago = (sale.get('MetodoPago', '') or 'efectivo').lower().strip()

                ingreso = (precio_venta * cantidad).quantize(Decimal("0.001"), ROUND_HALF_UP)
                costo = (costo_unitario * cantidad).quantize(Decimal("0.001"), ROUND_HALF_UP)
                utilidad = (ingreso - costo).quantize(Decimal("0.01"), ROUND_HALF_UP)

                utilidad = utilidad.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

                total_ingresos += ingreso
                total_costos += costo
                total_unidades += cantidad

                # Detalle de venta
                ventas_detalle.append({
                    'fecha': sale.get('Fecha', ''),
                    'hora': sale.get('Hora', ''),
                    'producto': sale.get('Nombre', ''),
                    'cantidad': float(cantidad),
                    'precio_venta': float(precio_venta),
                    'costo_unitario': float(costo_unitario),
                    'ingreso': float(ingreso),
                    'costo': float(costo),
                    'utilidad': float(utilidad),
                    'vendedor': vendedor
                })

                # Agrupar por producto
                if codigo not in productos_vendidos:
                    productos_vendidos[codigo] = {
                        'producto': sale.get('Nombre', ''),
                        'codigo': codigo,
                        'cantidad': Decimal("0"),
                        'ingresos': Decimal("0"),
                        'costos': Decimal("0"),
                        'utilidad': Decimal("0")
                    }
                productos_vendidos[codigo]['cantidad'] += cantidad
                productos_vendidos[codigo]['ingresos'] += ingreso
                productos_vendidos[codigo]['costos'] += costo
                productos_vendidos[codigo]['utilidad'] += utilidad

                # Estadísticas por vendedor
                if vendedor not in vendedores_stats:
                    vendedores_stats[vendedor] = {
                        'vendedor': vendedor,
                        'ventas': 0,
                        'ingresos': Decimal("0"),
                        'utilidad': Decimal("0")
                    }
                vendedores_stats[vendedor]['ventas'] += 1
                vendedores_stats[vendedor]['ingresos'] += ingreso
                vendedores_stats[vendedor]['utilidad'] += utilidad

                # Agrupar por método de pago
                if metodo_pago not in metodo_pago_stats:
                    metodo_pago_stats[metodo_pago] = {'metodo': metodo_pago, 'transacciones': 0, 'ingresos': Decimal("0")}
                metodo_pago_stats[metodo_pago]['transacciones'] += 1
                metodo_pago_stats[metodo_pago]['ingresos'] += ingreso
            
            utilidad_neta = (total_ingresos - total_costos).quantize(Decimal("0.01"), ROUND_HALF_UP)
            margen_total = ((utilidad_neta / total_ingresos * 100)).quantize(Decimal("0.01"), ROUND_HALF_UP)if total_ingresos > 0 else Decimal("0.00")
            
            # Convertir diccionarios a listas y ordenar
            productos_list = sorted(
                [{
                    **p,
                    'cantidad': float(p['cantidad']),
                    'ingresos': float(p['ingresos']),
                    'costos': float(p['costos']),
                    'utilidad': float(p['utilidad'])
                } for p in productos_vendidos.values()],
                key=lambda x: x['utilidad'],
                reverse=True
            )[:10]  # Top 10

            vendedores_list = sorted(
                [{
                    **v,
                    'ingresos': float(v['ingresos']),
                    'utilidad': float(v['utilidad'])
                } for v in vendedores_stats.values()],
                key=lambda x: x['ingresos'],
                reverse=True
            )

            metodo_pago_list = [
                {'metodo': m['metodo'], 'transacciones': m['transacciones'], 'ingresos': float(round(m['ingresos'], 2))}
                for m in sorted(metodo_pago_stats.values(), key=lambda x: x['ingresos'], reverse=True)
            ]

            print("Analis finalizado existosamente")
            return {
                'success': True,
                'data': {
                    'periodo': period_label,
                    'total_ingresos': float(round(total_ingresos, 2)),
                    'total_costos': float(round(total_costos, 2)),
                    'utilidad_neta': float(round(utilidad_neta, 2)),
                    'margen_total': float(round(margen_total, 2)),
                    'total_ventas': len(filtered_sales),
                    'total_unidades': float(total_unidades),
                    'ticket_promedio': float(round(total_ingresos / len(filtered_sales), 2)) if filtered_sales else 0,
                    'productos_vendidos': productos_list,
                    'vendedores': vendedores_list,
                    'ventas_detalle': ventas_detalle,
                    'metodo_pago_breakdown': metodo_pago_list,
                }
            }
            
        except Exception as e:
            import traceback
            return {
                'success': False,
                'error': str(e),
                'traceback': traceback.format_exc()
            }

    def export_cierre_caja(self, period='today', custom_start=None, custom_end=None):
        """Genera un archivo Excel con el cierre de caja del período dado"""
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        result = self.get_profit_analysis(period, custom_start, custom_end)
        if not result['success']:
            return result

        data = result['data']
        wb = openpyxl.Workbook()

        BLUE   = PatternFill(start_color='006BA6', end_color='006BA6', fill_type='solid')
        LBLUE  = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
        WHITE_BOLD = Font(color='FFFFFF', bold=True)
        CENTER = Alignment(horizontal='center', vertical='center')
        RIGHT  = Alignment(horizontal='right')

        def header_cell(ws, row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.font = WHITE_BOLD
            c.fill = BLUE
            c.alignment = CENTER
            return c

        def auto_width(ws):
            for col in ws.columns:
                width = max((len(str(cell.value or '')) for cell in col), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 45)

        # ── Sheet 1: Resumen ──────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = 'Resumen'

        ws1.merge_cells('A1:C1')
        t = ws1['A1']
        t.value = f"CIERRE DE CAJA — {data['periodo']}"
        t.font = Font(color='FFFFFF', bold=True, size=13)
        t.fill = BLUE
        t.alignment = CENTER
        ws1.row_dimensions[1].height = 28

        ws1.merge_cells('A2:C2')
        g = ws1['A2']
        g.value = f"Generado: {datetime.now(BUSINESS_TZ).strftime('%d/%m/%Y %H:%M:%S')}"
        g.font = Font(italic=True, color='666666', size=9)
        g.alignment = CENTER

        stats = [
            ('Ingresos Totales',  f"${data['total_ingresos']:.2f}"),
            ('Costos Totales',    f"${data['total_costos']:.2f}"),
            ('Utilidad Neta',     f"${data['utilidad_neta']:.2f}"),
            ('Margen',            f"{data['margen_total']:.2f}%"),
            ('Total de Ventas',   data['total_ventas']),
            ('Unidades Vendidas', data['total_unidades']),
            ('Ticket Promedio',   f"${data['ticket_promedio']:.2f}"),
        ]
        for i, (label, value) in enumerate(stats):
            r = i + 4
            lc = ws1.cell(row=r, column=1, value=label)
            lc.font = Font(bold=True)
            lc.fill = LBLUE
            vc = ws1.cell(row=r, column=2, value=value)
            vc.alignment = RIGHT

        if data['vendedores']:
            vstart = len(stats) + 6
            for col, h in enumerate(['Vendedor', 'Ventas', 'Ingresos', 'Utilidad'], 1):
                header_cell(ws1, vstart, col, h)
            for i, v in enumerate(data['vendedores']):
                r = vstart + 1 + i
                ws1.cell(row=r, column=1, value=v['vendedor'])
                ws1.cell(row=r, column=2, value=v['ventas'])
                ws1.cell(row=r, column=3, value=round(v['ingresos'], 2))
                ws1.cell(row=r, column=4, value=round(v['utilidad'], 2))

        auto_width(ws1)

        # ── Sheet 2: Detalle de Ventas ────────────────────────────────────
        ws2 = wb.create_sheet('Detalle de Ventas')
        cols2 = ['Fecha', 'Hora', 'Producto', 'Cantidad',
                 'Precio Venta', 'Costo Unitario', 'Ingreso', 'Costo', 'Utilidad', 'Vendedor']
        for col, h in enumerate(cols2, 1):
            header_cell(ws2, 1, col, h)

        for i, s in enumerate(data['ventas_detalle']):
            r = 2 + i
            for col, val in enumerate([
                s['fecha'], s['hora'], s['producto'], s['cantidad'],
                s['precio_venta'], s['costo_unitario'],
                s['ingreso'], s['costo'], s['utilidad'], s['vendedor']
            ], 1):
                ws2.cell(row=r, column=col, value=val)

        # Totals row
        if data['ventas_detalle']:
            tr = 2 + len(data['ventas_detalle'])
            tc = ws2.cell(row=tr, column=3, value='TOTAL')
            tc.font = Font(bold=True)
            for col, val in [(7, data['total_ingresos']), (8, data['total_costos']), (9, data['utilidad_neta'])]:
                c = ws2.cell(row=tr, column=col, value=round(val, 2))
                c.font = Font(bold=True)

        auto_width(ws2)

        # ── Sheet 3: Top Productos ────────────────────────────────────────
        ws3 = wb.create_sheet('Top Productos')
        cols3 = ['Producto', 'Código', 'Cantidad', 'Ingresos', 'Costos', 'Utilidad']
        for col, h in enumerate(cols3, 1):
            header_cell(ws3, 1, col, h)

        for i, p in enumerate(data['productos_vendidos']):
            r = 2 + i
            for col, val in enumerate([
                p['producto'], p['codigo'], p['cantidad'],
                round(p['ingresos'], 2), round(p['costos'], 2), round(p['utilidad'], 2)
            ], 1):
                ws3.cell(row=r, column=col, value=val)

        auto_width(ws3)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return {'success': True, 'workbook': output, 'periodo': data['periodo']}


class InferenceModel:
    """
    OCR-based product recognition from images of handwritten/printed sales lists.
    Uses PaddleOCR for text extraction and deterministic fuzzy matching against
    the inventory database — no external API calls, zero ongoing cost.
    """

    def __init__(self, _api_key=None):
        # api_key param kept for backward-compat but unused
        self._ocr = None  # Lazy init: PaddleOCR loads slowly on first use

    def _get_ocr(self):
        if self._ocr is None:
            from paddleocr import PaddleOCR
            self._ocr = PaddleOCR(use_angle_cls=True, lang='es', show_log=False)
        return self._ocr

    @staticmethod
    def _normalize(text):
        """Lowercase, strip accents, remove non-alphanumeric."""
        text = unicodedata.normalize('NFKD', str(text)).encode('ascii', 'ignore').decode('ascii')
        text = re.sub(r'[^a-z0-9\s]', ' ', text.lower())
        return re.sub(r'\s+', ' ', text).strip()

    @staticmethod
    def _levenshtein(s1, s2):
        """Normalized Levenshtein similarity [0.0 – 1.0]."""
        if not s1 or not s2:
            return 0.0
        n, m = len(s1), len(s2)
        prev = list(range(m + 1))
        for i in range(1, n + 1):
            curr = [i] + [0] * m
            for j in range(1, m + 1):
                curr[j] = prev[j - 1] if s1[i - 1] == s2[j - 1] else 1 + min(prev[j], curr[j - 1], prev[j - 1])
            prev = curr
        return 1.0 - prev[m] / max(n, m)

    def _match_product(self, token, inventory_items, threshold=0.65):
        """Return (item, score, method) for the best inventory match."""
        token_norm = self._normalize(token)
        if not token_norm:
            return None, 0.0, 'empty'

        best_item, best_score, best_method = None, 0.0, 'no_match'

        for item in inventory_items:
            codigo = str(item.get('Codigo', '')).strip()
            nombre = str(item.get('Nombre', ''))
            nombre_norm = self._normalize(nombre)

            # Exact code match — highest priority
            if token.strip().upper() == codigo.upper():
                return item, 1.0, 'exact_code'

            # Exact normalized name
            if token_norm == nombre_norm:
                return item, 0.97, 'exact_name'

            # Substring containment (partial word matching)
            if token_norm and nombre_norm:
                if token_norm in nombre_norm:
                    score = len(token_norm) / len(nombre_norm)
                    score = min(0.90, max(0.70, score))
                elif nombre_norm in token_norm:
                    score = len(nombre_norm) / len(token_norm)
                    score = min(0.88, max(0.65, score))
                else:
                    score = self._levenshtein(token_norm, nombre_norm)

                if score > best_score and score >= threshold:
                    best_score, best_item, best_method = score, item, 'fuzzy'

        return best_item, best_score, best_method

    def infer_cart_from_image(self, image_file, inventory_items=None):
        """
        Extracts a cart from an image of a sales list.
        Steps:
          1. PaddleOCR extracts text lines
          2. Regex heuristics parse quantity + product token per line
          3. Fuzzy match each token against inventory
        Returns list of cart dicts compatible with process_sale().
        """
        import io

        try:
            image_bytes = image_file.read() if hasattr(image_file, 'read') else image_file
        except Exception:
            image_bytes = image_file

        ocr = self._get_ocr()
        result = ocr.ocr(image_bytes, cls=True)

        if not result or not result[0]:
            print("OCR returned no results")
            return []

        lines = [(line[1][0].strip(), float(line[1][1])) for line in result[0] if line[1][0].strip()]
        print(f"OCR extracted {len(lines)} lines: {[l[0] for l in lines]}")

        # Quantity extraction patterns (qty first or qty last)
        qty_first_re  = re.compile(r'^(\d+(?:[.,]\d+)?)\s*[xX\*\-]?\s+(.+)', re.UNICODE)
        qty_last_re   = re.compile(r'^(.+?)\s+[xX\*]\s*(\d+(?:[.,]\d+)?)$', re.UNICODE)
        qty_suffix_re = re.compile(r'^(.+?)\s+(\d+(?:[.,]\d+)?)$', re.UNICODE)

        cart = []
        seen_codes = {}  # code -> cart index for duplicate accumulation

        for raw_text, ocr_conf in lines:
            quantity = 1.0
            product_token = raw_text

            # Try to extract quantity
            for pat, qty_group, tok_group in [
                (qty_first_re,  1, 2),
                (qty_last_re,   2, 1),
                (qty_suffix_re, 2, 1),
            ]:
                m = pat.match(raw_text)
                if m:
                    try:
                        q = float(m.group(qty_group).replace(',', '.'))
                        if 0 < q <= 9999:
                            quantity = q
                            product_token = m.group(tok_group).strip()
                            break
                    except ValueError:
                        pass

            if not inventory_items or not product_token:
                continue

            matched, score, method = self._match_product(product_token, inventory_items)

            if matched and score >= 0.65:
                codigo = str(matched.get('Codigo', '')).strip()

                if codigo in seen_codes:
                    cart[seen_codes[codigo]]['cantidadVendida'] += quantity
                    continue

                seen_codes[codigo] = len(cart)
                cart.append({
                    'codigo': codigo,
                    'nombre': str(matched.get('Nombre', '')),
                    'cantidadVendida': quantity,
                    'precio': float(str(matched.get('Precio_1', 0)) or 0),
                    'tipoPrecio': 'precio',
                    'confidence': round(score, 2),
                    'original_text': raw_text,
                    'match_method': method,
                })
            else:
                # Include unmatched lines so the user can see what OCR read
                cart.append({
                    'codigo': '',
                    'nombre': product_token,
                    'cantidadVendida': quantity,
                    'precio': 0,
                    'tipoPrecio': 'precio',
                    'confidence': 0.0,
                    'original_text': raw_text,
                    'match_method': 'no_match',
                })

        print(f"Cart built: {len(cart)} items, {len([c for c in cart if c['confidence'] >= 0.65])} matched")
        return cart

# Ejemplo de uso
if __name__ == "__main__":
    # Inicializar
    inventory = InventoryManager(
        'credenciales.json',
        'Inventario_MiTienda'
    )
    
    # Ejemplo de venta
    venta = [
        {'codigo': 'CAM001', 'cantidad_vendida': 2},
        {'codigo': 'PAN001', 'cantidad_vendida': 1}
    ]
    
    resultado = inventory.process_sale(venta)
    
    if resultado['success']:
        print("Venta procesada exitosamente")
        
        if resultado['alerts']:
            print("\nALERTAS DE STOCK BAJO:")
            for alert in resultado['alerts']:
                print(f"  - {alert['producto']}: {alert['cantidad_restante']} unidades")
    else:
        print(" Error procesando la venta")