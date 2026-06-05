import gspread
from oauth2client.service_account import ServiceAccountCredentials
import hashlib
from datetime import datetime
from zoneinfo import ZoneInfo
import uuid
import time
from decimal import Decimal, ROUND_HALF_UP
from babel.dates import format_date

BUSINESS_TZ = ZoneInfo("America/Guayaquil")


class InventoryManager:
    def __init__(self, credentials_file, spreadsheet_name):
        scope = ['https://spreadsheets.google.com/feeds',
                 'https://www.googleapis.com/auth/drive']

        creds = ServiceAccountCredentials.from_json_keyfile_name(
            credentials_file, scope
        )
        self.client = gspread.authorize(creds)
        self.spreadsheet = self.client.open(spreadsheet_name)
        self.sheet_inventory = self.spreadsheet.worksheet('Inventario')
        self.sheet_sales = self.spreadsheet.worksheet('Ventas')
        self.sheet_users = self.spreadsheet.worksheet('Usuarios')
        self._row_cache = {}
        self._row_cache_timestamp = 0
        self._row_cache_ttl = 300

    def _get_product_row(self, product_code):
        """Get row number from cache or fetch from sheet"""
        now = time.time()

        if (now - self._row_cache_timestamp) > self._row_cache_ttl:
            print("Refreshing row cache...")
            codes = self.sheet_inventory.col_values(2)
            self._row_cache = {code: i + 1 for i, code in enumerate(codes) if code}
            self._row_cache_timestamp = now
            print(f"Cache loaded with {len(self._row_cache)} products")

        return self._row_cache.get(product_code)

    def _get_or_create_sheet(self, title, headers):
        """Gets or creates a worksheet with the given title and headers"""
        try:
            return self.spreadsheet.worksheet(title)
        except Exception:
            sheet = self.spreadsheet.add_worksheet(title=title, rows=1000, cols=len(headers))
            sheet.append_row(headers)
            return sheet
        
    def hash_password(self, password):
        """Hash de contraseña con SHA256"""
        return hashlib.sha256(password.encode()).hexdigest()

    def create_user(self, username, password, role='vendedor', nombre=''):
        """Crea un nuevo usuario"""
        try:
            users = self.sheet_users.get_all_records()
            for user in users:
                if user['Usuario'].lower() == username.lower():
                    return {'success': False, 'message': 'El usuario ya existe'}

            hashed_password = self.hash_password(password)
            next_id = len(users) + 1
            row = [next_id, username, hashed_password, role, nombre, 'Si', '']
            self.sheet_users.append_row(row)

            return {'success': True, 'message': 'Usuario creado exitosamente', 'user_id': next_id}

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def authenticate_user(self, username, password):
        """Autentica un usuario"""
        try:
            users = self.sheet_users.get_all_records()
            hashed_password = self.hash_password(password)

            for user in users:
                if (user['Usuario'].lower() == username.lower() and
                        user['Password'] == hashed_password and
                        user['Activo'].lower() == 'si'):

                    user_row = users.index(user) + 2
                    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    self.sheet_users.update_cell(user_row, 7, now)

                    return {
                        'success': True,
                        'user': {
                            'id': user['ID'],
                            'username': user['Usuario'],
                            'role': user['Rol'],
                            'nombre': user['Nombre']
                        }
                    }

            return {'success': False, 'message': 'Usuario o contraseña incorrectos'}

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def get_all_users(self):
        """Obtiene todos los usuarios (sin passwords)"""
        try:
            users = self.sheet_users.get_all_records()
            users_list = []

            for user in users:
                users_list.append({
                    'id': user['ID'],
                    'username': user['Usuario'],
                    'role': user['Rol'],
                    'nombre': user['Nombre'],
                    'activo': user['Activo'],
                    'ultimo_acceso': user.get('UltimoAcceso', '')
                })

            return {'success': True, 'users': users_list}

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def get_inventory(self):
        """Obtiene todo el inventario"""
        records = self.sheet_inventory.get_all_records()
        return records

    def add_product(self, product_data):
        """Agrega un nuevo producto a la hoja de Inventario"""
        try:
            now = datetime.now(BUSINESS_TZ)
            ultima_actualizacion = now.strftime('%Y-%m-%d %H:%M:%S')

            all_records = self.sheet_inventory.get_all_records()
            next_id = len(all_records) + 1
            new_row = len(all_records) + 2

            row = [
                next_id,
                product_data['codigo'],
                product_data['nombre'],
                product_data['cantidad'],
                product_data['unidad'],
                product_data['costo'],
                product_data['precio_1'],
                product_data['precio_2'],
                product_data['precio_3'],
                product_data['minStock'],
                ultima_actualizacion,
                product_data.get('categoria', ''),
                product_data.get('subcategoria', ''),
                float(product_data.get('descuento', 0) or 0),
            ]

            print(f'Producto para insertar: {row}')
            self.sheet_inventory.append_row(row)

            self._row_cache[product_data['codigo']] = new_row
            print(f"Cache actualizado: {product_data['codigo']} -> row {new_row}")

            return {
                'success': True,
                'message': 'Producto agregado exitosamente',
                'product_code': product_data['codigo'],
                'product_id': next_id
            }

        except Exception as e:
            return {'success': False, 'error': str(e), 'message': f'Error al agregar producto: {str(e)}'}

    def update_product(self, product_code, updates):
        """Actualiza los campos de un producto existente (sin tocar cantidad)"""
        try:
            row = self._get_product_row(product_code)
            if row is None:
                return {'success': False, 'error': f'Producto {product_code} no encontrado'}

            now = datetime.now(BUSINESS_TZ).strftime('%Y-%m-%d %H:%M:%S')
            batch = []
            if 'nombre' in updates:
                batch.append({'range': f'C{row}', 'values': [[updates['nombre']]]})
            if 'costo' in updates:
                batch.append({'range': f'F{row}', 'values': [[updates['costo']]]})
            if 'precio_1' in updates:
                batch.append({'range': f'G{row}', 'values': [[updates['precio_1']]]})
            if 'precio_2' in updates:
                batch.append({'range': f'H{row}', 'values': [[updates['precio_2']]]})
            if 'precio_3' in updates:
                batch.append({'range': f'I{row}', 'values': [[updates['precio_3']]]})
            if 'minStock' in updates:
                batch.append({'range': f'J{row}', 'values': [[updates['minStock']]]})
            if 'categoria' in updates:
                batch.append({'range': f'L{row}', 'values': [[updates['categoria']]]})
            if 'subcategoria' in updates:
                batch.append({'range': f'M{row}', 'values': [[updates['subcategoria']]]})
            if 'descuento' in updates:
                batch.append({'range': f'N{row}', 'values': [[float(updates['descuento'] or 0)]]})
            batch.append({'range': f'K{row}', 'values': [[now]]})

            if batch:
                self.sheet_inventory.batch_update(batch)

            self._row_cache_timestamp = 0
            return {'success': True, 'message': 'Producto actualizado correctamente'}

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def adjust_stock(self, product_code, cantidad_ajuste, motivo=''):
        """Ajusta el stock de un producto (puede ser positivo o negativo)"""
        try:
            row = self._get_product_row(product_code)
            if row is None:
                return {'success': False, 'error': f'Producto {product_code} no encontrado'}

            row_values = self.sheet_inventory.row_values(row)
            current_qty = float(row_values[3])
            new_qty = round(current_qty + float(cantidad_ajuste), 2)

            if new_qty < 0:
                return {'success': False, 'error': f'El ajuste dejaría el stock en {new_qty}, no puede ser negativo'}

            now = datetime.now(BUSINESS_TZ).strftime('%Y-%m-%d %H:%M:%S')
            self.sheet_inventory.batch_update([
                {'range': f'D{row}', 'values': [[new_qty]]},
                {'range': f'K{row}', 'values': [[now]]}
            ])

            try:
                ajustes_sheet = self._get_or_create_sheet('Ajustes', ['Fecha', 'Codigo', 'Nombre', 'AjusteAnterior', 'Ajuste', 'NuevoStock', 'Motivo'])
                ajustes_sheet.append_row([now, product_code, row_values[2], current_qty, cantidad_ajuste, new_qty, motivo])
            except Exception:
                pass  # log failure is non-critical

            self._row_cache_timestamp = 0
            return {'success': True, 'new_quantity': new_qty, 'previous_quantity': current_qty}

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def get_product_by_code(self, code):
        """Busca un producto por código"""
        try:
            cell = self.sheet_inventory.find(code)
            row = self.sheet_inventory.row_values(cell.row)
            return {
                'id': row[0],
                'codigo': row[1],
                'nombre': row[2],
                'cantidad': int(row[3]),
                'precio': float(row[4]),
                'minStock': int(row[5])
            }
        except Exception:
            return None

    def update_stock(self, product_code, quantity_sold, price_type):
        """Actualiza el stock después de una venta"""
        try:
            print(f"Updating {product_code}")
            row = self._get_product_row(product_code)

            if row is None:
                return {'success': False, 'error': f'Producto {product_code} no encontrado'}

            row_values = self.sheet_inventory.row_values(row)

            product_id = row_values[0]
            product_name = row_values[2]
            current_qty = float(row_values[3])
            unidad = row_values[4].lower()
            price_1 = float(row_values[6])
            price_2 = float(row_values[7]) if row_values[7] else 0.0
            price_3 = float(row_values[8]) if row_values[8] else 0.0
            min_stock = float(row_values[9])
            descuento = float(row_values[13]) if len(row_values) > 13 and row_values[13] else 0.0
            quantity_sold = float(quantity_sold)

            if unidad == "unidad" and not quantity_sold.is_integer():
                return {'success': False, 'error': 'Este producto solo se puede vender en unidades enteras'}

            if current_qty < quantity_sold:
                return {'success': False, 'error': 'Stock insuficiente'}

            new_qty = round(current_qty - quantity_sold, 2)
            timestamp = datetime.now(BUSINESS_TZ).strftime('%Y-%m-%d %H:%M:%S')

            self.sheet_inventory.batch_update([
                {'range': f'D{row}', 'values': [[new_qty]]},
                {'range': f'K{row}', 'values': [[timestamp]]}
            ])

            alert = new_qty <= min_stock

            if price_type == "precio_2":
                selected_price = price_2
            elif price_type == "precio_3":
                selected_price = price_3
            else:
                selected_price = price_1

            if descuento > 0:
                selected_price = round(selected_price * (1 - descuento / 100), 4)

            return {
                'success': True,
                'product_id': product_id,
                'product_code': product_code,
                'product_name': product_name,
                'price': selected_price,
                'descuento': descuento,
                'quantity_sold': quantity_sold,
                'new_quantity': new_qty,
                'alert': alert
            }

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def save_sale(self, sale_id, sale_details, total, vendedor='Sistema', metodo_pago='efectivo', referencia=''):
        """Guarda el detalle de la venta en la hoja de Ventas"""
        try:
            now = datetime.now(BUSINESS_TZ)
            fecha = now.strftime('%Y-%m-%d')
            hora = now.strftime('%H:%M:%S')

            rows = []
            for item in sale_details:
                row = [
                    sale_id, fecha, hora,
                    item['product_id'], item['product_code'], item['product_name'],
                    item['quantity_sold'], item['price'],
                    round(item['price'] * item['quantity_sold'], 2),
                    round(total, 2),
                    vendedor, metodo_pago, referencia,
                    item.get('descuento', 0),
                ]
                rows.append(row)

            print(f'Venta para insertar: {rows}')
            self.sheet_sales.append_rows(rows, value_input_option='USER_ENTERED')
            print(f'Venta guardada')

            return {'success': True, 'sale_id': sale_id, 'items_saved': len(rows)}

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def process_sale(self, cart_items, vendedor='Sistema', metodo_pago='efectivo', referencia=''):
        """Procesa una venta completa"""
        sale_id = f"VTA-{datetime.now(BUSINESS_TZ).strftime('%Y%m%d')}-{str(uuid.uuid4())[:8]}"

        results = []
        alerts = []
        total_sale = 0
        sale_details = []

        for item in cart_items:
            result = self.update_stock(item['codigo'], item['cantidad_vendida'], item['tipoPrecio'])

            if not result['success']:
                return {'success': False, 'error': f"Error en {item['codigo']}: {result['error']}"}

            results.append(result)

            subtotal = result['price'] * result['quantity_sold']
            total_sale += subtotal

            sale_details.append({
                'product_id': result['product_id'],
                'product_code': result['product_code'],
                'product_name': result['product_name'],
                'price': result['price'],
                'descuento': result['descuento'],
                'quantity_sold': result['quantity_sold']
            })

            if result.get('alert'):
                alerts.append({'producto': result['product_name'], 'cantidad_restante': result['new_quantity']})

        print("Inventario actualizado")

        total_con_descuento = round(total_sale, 2)
        save_result = self.save_sale(sale_id, sale_details, total_con_descuento, vendedor, metodo_pago, referencia)

        if not save_result['success']:
            return {'success': False, 'error': f"Venta procesada pero no se guardó en historial: {save_result['error']}"}

        return {
            'success': True,
            'sale_id': sale_id,
            'total': total_con_descuento,
            'subtotal': total_sale,
            'descuento': 0,
            'items': len(results),
            'results': results,
            'alerts': alerts
        }

    def refund_sale(self, sale_id, items):
        """
        Process a refund for a sale.
        items: list of {codigo, cantidad} dicts.
        Restores stock for each item and writes negative rows to Ventas sheet.
        """
        try:
            all_sales = self.sheet_sales.get_all_records()
            original_rows = [r for r in all_sales if str(r.get('VentaID', '')) == sale_id]

            if not original_rows:
                return {'success': False, 'error': f'Venta {sale_id} no encontrada'}

            price_lookup = {str(r['Codigo']): float(r.get('PrecioUnitario', 0)) for r in original_rows}

            refund_id = f"DEV-{datetime.now(BUSINESS_TZ).strftime('%Y%m%d')}-{str(uuid.uuid4())[:8]}"
            now = datetime.now(BUSINESS_TZ)
            fecha = now.strftime('%Y-%m-%d')
            hora = now.strftime('%H:%M:%S')

            refund_rows = []
            total_refund = 0

            for item in items:
                codigo = str(item.get('codigo', '')).strip()
                cantidad = float(item.get('cantidad', 0))

                if not codigo or cantidad <= 0:
                    continue

                row = self._get_product_row(codigo)
                if row is None:
                    return {'success': False, 'error': f'Producto {codigo} no encontrado en inventario'}

                row_values = self.sheet_inventory.row_values(row)
                current_qty = float(row_values[3])
                new_qty = round(current_qty + cantidad, 2)
                timestamp = datetime.now(BUSINESS_TZ).strftime('%Y-%m-%d %H:%M:%S')
                self.sheet_inventory.batch_update([
                    {'range': f'D{row}', 'values': [[new_qty]]},
                    {'range': f'K{row}', 'values': [[timestamp]]}
                ])

                precio = price_lookup.get(codigo, 0)
                subtotal = round(-precio * cantidad, 2)
                total_refund += subtotal

                refund_rows.append([
                    refund_id, fecha, hora,
                    row_values[0], codigo, row_values[2],
                    -cantidad, precio, subtotal,
                    round(total_refund, 2),
                    'DEVOLUCION', 'devolucion', sale_id, 0,
                ])

            if refund_rows:
                self.sheet_sales.append_rows(refund_rows, value_input_option='USER_ENTERED')

            self._row_cache_timestamp = 0

            return {
                'success': True,
                'refund_id': refund_id,
                'items_refunded': len(refund_rows),
                'total_refund': abs(total_refund),
            }

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def get_sales_history(self, limit=None, date_from=None, date_to=None):
        """Obtiene el historial de ventas con filtros opcionales"""
        try:
            records = self.sheet_sales.get_all_records()

            if date_from:
                records = [r for r in records if r['Fecha'] >= date_from]
            if date_to:
                records = [r for r in records if r['Fecha'] <= date_to]
            if limit:
                records = records[-limit:]

            return records

        except Exception as e:
            print(f"Error obteniendo historial: {e}")
            return []

    def get_sales_summary(self, date=None):
        """Obtiene un resumen de ventas del día"""
        try:
            if date is None:
                date = datetime.now(BUSINESS_TZ).strftime('%Y-%m-%d')

            records = self.sheet_sales.get_all_records()
            daily_sales = [r for r in records if r['Fecha'] == date]

            if not daily_sales:
                return {'date': date, 'total_sales': 0, 'total_amount': 0, 'items_sold': 0, 'unique_sales': 0}

            unique_sales = len(set(r['VentaID'] for r in daily_sales))
            total_items = sum(r['Cantidad'] for r in daily_sales)
            total_amount = sum(r['Subtotal'] for r in daily_sales)

            return {
                'date': date,
                'total_sales': unique_sales,
                'total_amount': total_amount,
                'items_sold': total_items,
                'sales': daily_sales
            }

        except Exception as e:
            return {'error': str(e)}

    def get_sales_chart(self, period='today'):
        """Returns time-series sales data for charts (revenue + sales count per bucket)"""
        from datetime import datetime, timedelta

        now = datetime.now(BUSINESS_TZ)

        if period == 'today':
            date_str = now.strftime('%Y-%m-%d')
            records = self.get_sales_history(date_from=date_str, date_to=date_str)
            buckets = {str(h).zfill(2): {'revenue': 0.0, 'sales': set()} for h in range(24)}
            for r in records:
                hora = str(r.get('Hora', ''))
                h = hora.split(':')[0].zfill(2) if hora else None
                if h and h in buckets:
                    buckets[h]['revenue'] += float(r.get('Subtotal', 0) or 0)
                    buckets[h]['sales'].add(str(r.get('VentaID', '')))
            labels = [f"{h}:00" for h in range(24)]
            revenue = [round(buckets[str(h).zfill(2)]['revenue'], 2) for h in range(24)]
            counts = [len(buckets[str(h).zfill(2)]['sales']) for h in range(24)]

        elif period == 'week':
            start = now - timedelta(days=now.weekday())
            date_from = start.strftime('%Y-%m-%d')
            date_to = now.strftime('%Y-%m-%d')
            records = self.get_sales_history(date_from=date_from, date_to=date_to)
            days_count = now.weekday() + 1
            day_dates = [(start + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(days_count)]
            day_names = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
            buckets = {d: {'revenue': 0.0, 'sales': set()} for d in day_dates}
            for r in records:
                d = str(r.get('Fecha', ''))
                if d in buckets:
                    buckets[d]['revenue'] += float(r.get('Subtotal', 0) or 0)
                    buckets[d]['sales'].add(str(r.get('VentaID', '')))
            labels = [day_names[(start + timedelta(days=i)).weekday()] for i in range(days_count)]
            revenue = [round(buckets[d]['revenue'], 2) for d in day_dates]
            counts = [len(buckets[d]['sales']) for d in day_dates]

        elif period == 'month':
            start = now.replace(day=1)
            date_from = start.strftime('%Y-%m-%d')
            date_to = now.strftime('%Y-%m-%d')
            records = self.get_sales_history(date_from=date_from, date_to=date_to)
            days_count = now.day
            day_dates = [(start + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(days_count)]
            buckets = {d: {'revenue': 0.0, 'sales': set()} for d in day_dates}
            for r in records:
                d = str(r.get('Fecha', ''))
                if d in buckets:
                    buckets[d]['revenue'] += float(r.get('Subtotal', 0) or 0)
                    buckets[d]['sales'].add(str(r.get('VentaID', '')))
            labels = [(start + timedelta(days=i)).strftime('%d/%m') for i in range(days_count)]
            revenue = [round(buckets[d]['revenue'], 2) for d in day_dates]
            counts = [len(buckets[d]['sales']) for d in day_dates]

        else:
            return {'success': False, 'error': 'Período no válido'}

        total_revenue = round(sum(revenue), 2)
        all_sale_ids = set(str(r.get('VentaID', '')) for r in records if r.get('VentaID'))
        total_sales = len(all_sale_ids)
        avg_ticket = round(total_revenue / total_sales, 2) if total_sales > 0 else 0

        return {
            'success': True,
            'data': {
                'labels': labels,
                'revenue': revenue,
                'sales_count': counts,
                'total_revenue': total_revenue,
                'total_sales': total_sales,
                'avg_ticket': avg_ticket,
            }
        }

    def _get_or_create_categories_sheet(self):
        try:
            return self.spreadsheet.worksheet('Categorias')
        except gspread.exceptions.WorksheetNotFound:
            sheet = self.spreadsheet.add_worksheet(title='Categorias', rows=200, cols=2)
            sheet.append_row(['Categoria', 'Subcategoria'])
            return sheet

    def get_categories(self):
        """Obtiene categorías y subcategorías desde la hoja Categorias"""
        try:
            sheet = self._get_or_create_categories_sheet()
            records = sheet.get_all_records()
            grouped = {}
            for row in records:
                cat = str(row.get('Categoria', '')).strip()
                sub = str(row.get('Subcategoria', '')).strip()
                if not cat:
                    continue
                if cat not in grouped:
                    grouped[cat] = []
                if sub and sub not in grouped[cat]:
                    grouped[cat].append(sub)
            return {'success': True, 'data': grouped}

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def add_category(self, categoria, subcategoria=''):
        """Agrega una categoría o subcategoría a la hoja Categorias"""
        try:
            categoria = categoria.strip()
            subcategoria = subcategoria.strip() if subcategoria else ''
            if not categoria:
                return {'success': False, 'error': 'El nombre de la categoría no puede estar vacío'}
            sheet = self._get_or_create_categories_sheet()
            records = sheet.get_all_records()
            for row in records:
                if (str(row.get('Categoria', '')).strip() == categoria and
                        str(row.get('Subcategoria', '')).strip() == subcategoria):
                    return {'success': False, 'error': 'Esta categoría/subcategoría ya existe'}
            sheet.append_row([categoria, subcategoria])
            return {'success': True, 'message': 'Categoría agregada exitosamente'}

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def save_invoice_to_sheet(self, sale_id, invoice_result, cliente, cart_items, total):
        """Guarda los datos de una factura electrónica en la hoja 'Facturas'."""
        try:
            headers = [
                'FacturaID', 'VentaID', 'Fecha', 'Hora',
                'ClaveAcceso', 'NumeroFactura', 'NumAutorizacion', 'FechaAutorizacion',
                'ClienteID', 'ClienteNombre', 'ClienteEmail',
                'Total', 'Ambiente',
            ]
            sheet = self._get_or_create_sheet('Facturas', headers)
            now = datetime.now(BUSINESS_TZ)
            row = [
                f"FAC-{now.strftime('%Y%m%d%H%M%S')}",
                sale_id,
                now.strftime('%Y-%m-%d'),
                now.strftime('%H:%M:%S'),
                invoice_result.get('clave_acceso', ''),
                invoice_result.get('numero_factura', ''),
                invoice_result.get('numero_autorizacion', ''),
                invoice_result.get('fecha_autorizacion', ''),
                cliente.get('identificacion', ''),
                cliente.get('razon_social', ''),
                cliente.get('email', ''),
                total,
                invoice_result.get('ambiente', ''),
            ]
            sheet.append_row(row, value_input_option='USER_ENTERED')
            return {'success': True}

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def send_invoice_email(self, cliente_email, cliente_nombre, invoice_result, xml_path=None):
        """Envía la factura autorizada al cliente por correo electrónico."""
        import smtplib
        import os
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders

        smtp_host = os.environ.get('SMTP_HOST', '')
        smtp_port = int(os.environ.get('SMTP_PORT', 587))
        smtp_user = os.environ.get('SMTP_USER', '')
        smtp_pass = os.environ.get('SMTP_PASS', '')

        if not smtp_host or not smtp_user or not smtp_pass:
            return {'success': False, 'error': 'Credenciales SMTP no configuradas (SMTP_HOST, SMTP_USER, SMTP_PASS)'}

        try:
            msg = MIMEMultipart()
            msg['From'] = smtp_user
            msg['To'] = cliente_email
            msg['Subject'] = f"Factura Electrónica – {invoice_result.get('numero_factura', '')}"

            body = (
                f"Estimado/a {cliente_nombre},\n\n"
                f"Adjuntamos su factura electrónica autorizada por el SRI.\n\n"
                f"Número: {invoice_result.get('numero_factura', '')}\n"
                f"Clave de acceso: {invoice_result.get('clave_acceso', '')}\n"
                f"Autorización: {invoice_result.get('numero_autorizacion', '')}\n"
                f"Fecha: {invoice_result.get('fecha_autorizacion', '')}\n\n"
                "Saludos,\nCentro Comercial TB"
            )
            msg.attach(MIMEText(body, 'plain', 'utf-8'))

            if xml_path and os.path.isfile(xml_path):
                with open(xml_path, 'rb') as f:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(f.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(xml_path)}"')
                    msg.attach(part)

            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, cliente_email, msg.as_string())

            return {'success': True}

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def get_low_stock_alerts(self):
        """Obtiene todos los productos con stock bajo"""
        records = self.get_inventory()
        alerts = []

        for record in records:
            if record['Cantidad'] <= record['MinStock']:
                alerts.append({
                    'codigo': record['Codigo'],
                    'nombre': record['Nombre'],
                    'cantidad': record['Cantidad'],
                    'minimo': record['MinStock']
                })

        return alerts

    def get_profit_analysis(self, period='today', custom_start=None, custom_end=None):
        """Analiza las utilidades para cierre de caja por período"""
        try:
            from datetime import datetime, timedelta

            now = datetime.now(BUSINESS_TZ)

            if period == 'today':
                print("Filtrando ventas de hoy")
                start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
                end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
                period_label = f"Hoy - {now.strftime('%d/%m/%Y')}"

            elif period == 'week':
                print("Filtrando ventas de la semana")
                start_date = now - timedelta(days=now.weekday())
                start_date = start_date.replace(hour=0, minute=0, second=0)
                end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
                period_label = f"Esta Semana ({start_date.strftime('%d/%m')} - {end_date.strftime('%d/%m/%Y')})"

            elif period == 'month':
                print("Filtrando ventas del mes")
                start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
                period_label = f"Este Mes - {now.strftime('%B %Y')}"

            elif period == 'custom' and custom_start and custom_end:
                print("Filtrando ventas en rango personalizado")
                start_date = datetime.strptime(custom_start, '%Y-%m-%d').replace(tzinfo=BUSINESS_TZ)
                end_date = datetime.strptime(custom_end, '%Y-%m-%d').replace(hour=23, minute=59, second=59, tzinfo=BUSINESS_TZ)
            else:
                return {'success': False, 'error': 'Período no válido'}

            date_col = self.sheet_sales.col_values(2)[1:]
            headers = self.sheet_sales.row_values(1)

            matching_rows = []
            for i, date_str in enumerate(date_col):
                try:
                    sale_date = datetime.strptime(date_str, '%Y-%m-%d').replace(tzinfo=BUSINESS_TZ)
                    if start_date.date() <= sale_date.date() <= end_date.date():
                        matching_rows.append(i + 2)
                except Exception:
                    continue

            if not matching_rows:
                return {
                    'success': True,
                    'data': {
                        'periodo': period_label,
                        'total_ingresos': 0, 'total_costos': 0,
                        'utilidad_neta': 0, 'margen_total': 0,
                        'total_ventas': 0, 'total_unidades': 0,
                        'ticket_promedio': 0,
                        'productos_vendidos': [], 'vendedores': [],
                        'ventas_detalle': [], 'metodo_pago_breakdown': [],
                    }
                }

            col_letter = chr(64 + len(headers))
            sheet_name = self.sheet_sales.title
            ranges = [f"'{sheet_name}'!A{row}:{col_letter}{row}" for row in matching_rows]

            batch_data = self.sheet_sales.spreadsheet.values_batch_get(ranges)

            filtered_sales = []
            for value_range in batch_data.get('valueRanges', []):
                values = value_range.get('values', [])
                if values:
                    record = dict(zip(headers, values[0]))
                    filtered_sales.append(record)

            inventory = self.sheet_inventory.get_all_records()
            costs_dict = {item['Codigo']: float(item.get('Costo', 0)) for item in inventory}

            total_ingresos = Decimal("0.000")
            total_costos = Decimal("0.000")
            total_unidades = Decimal("0")
            ventas_detalle = []
            productos_vendidos = {}
            vendedores_stats = {}
            metodo_pago_stats = {}

            for sale in filtered_sales:
                codigo = sale.get('Codigo', '')
                cantidad = Decimal(str(sale.get('Cantidad', 0)))
                precio_venta = Decimal(str(sale.get('PrecioUnitario', 0)))
                costo_unitario = Decimal(str(costs_dict.get(codigo, 0)))
                vendedor = sale.get('Vendedor', 'Sistema')
                metodo_pago = (sale.get('MetodoPago', '') or 'efectivo').lower().strip()

                ingreso = (precio_venta * cantidad).quantize(Decimal("0.001"), ROUND_HALF_UP)
                costo = (costo_unitario * cantidad).quantize(Decimal("0.001"), ROUND_HALF_UP)
                utilidad = (ingreso - costo).quantize(Decimal("0.01"), ROUND_HALF_UP)

                total_ingresos += ingreso
                total_costos += costo
                total_unidades += cantidad

                ventas_detalle.append({
                    'fecha': sale.get('Fecha', ''),
                    'hora': sale.get('Hora', ''),
                    'producto': sale.get('Nombre', ''),
                    'cantidad': float(cantidad),
                    'precio_venta': float(precio_venta),
                    'costo_unitario': float(costo_unitario),
                    'ingreso': float(ingreso),
                    'costo': float(costo),
                    'utilidad': float(utilidad),
                    'vendedor': vendedor
                })

                if codigo not in productos_vendidos:
                    productos_vendidos[codigo] = {
                        'producto': sale.get('Nombre', ''),
                        'codigo': codigo,
                        'cantidad': Decimal("0"),
                        'ingresos': Decimal("0"),
                        'costos': Decimal("0"),
                        'utilidad': Decimal("0")
                    }
                productos_vendidos[codigo]['cantidad'] += cantidad
                productos_vendidos[codigo]['ingresos'] += ingreso
                productos_vendidos[codigo]['costos'] += costo
                productos_vendidos[codigo]['utilidad'] += utilidad

                if vendedor not in vendedores_stats:
                    vendedores_stats[vendedor] = {'vendedor': vendedor, 'ventas': 0, 'ingresos': Decimal("0"), 'utilidad': Decimal("0")}
                vendedores_stats[vendedor]['ventas'] += 1
                vendedores_stats[vendedor]['ingresos'] += ingreso
                vendedores_stats[vendedor]['utilidad'] += utilidad

                if metodo_pago not in metodo_pago_stats:
                    metodo_pago_stats[metodo_pago] = {'metodo': metodo_pago, 'transacciones': 0, 'ingresos': Decimal("0")}
                metodo_pago_stats[metodo_pago]['transacciones'] += 1
                metodo_pago_stats[metodo_pago]['ingresos'] += ingreso

            utilidad_neta = (total_ingresos - total_costos).quantize(Decimal("0.01"), ROUND_HALF_UP)
            margen_total = (utilidad_neta / total_ingresos * 100).quantize(Decimal("0.01"), ROUND_HALF_UP) if total_ingresos > 0 else Decimal("0.00")

            productos_list = sorted(
                [{**p, 'cantidad': float(p['cantidad']), 'ingresos': float(p['ingresos']),
                  'costos': float(p['costos']), 'utilidad': float(p['utilidad'])}
                 for p in productos_vendidos.values()],
                key=lambda x: x['utilidad'], reverse=True
            )[:10]

            vendedores_list = sorted(
                [{**v, 'ingresos': float(v['ingresos']), 'utilidad': float(v['utilidad'])}
                 for v in vendedores_stats.values()],
                key=lambda x: x['ingresos'], reverse=True
            )

            metodo_pago_list = [
                {'metodo': m['metodo'], 'transacciones': m['transacciones'], 'ingresos': float(round(m['ingresos'], 2))}
                for m in sorted(metodo_pago_stats.values(), key=lambda x: x['ingresos'], reverse=True)
            ]

            print("Análisis finalizado exitosamente")
            return {
                'success': True,
                'data': {
                    'periodo': period_label,
                    'total_ingresos': float(round(total_ingresos, 2)),
                    'total_costos': float(round(total_costos, 2)),
                    'utilidad_neta': float(round(utilidad_neta, 2)),
                    'margen_total': float(round(margen_total, 2)),
                    'total_ventas': len(filtered_sales),
                    'total_unidades': float(total_unidades),
                    'ticket_promedio': float(round(total_ingresos / len(filtered_sales), 2)) if filtered_sales else 0,
                    'productos_vendidos': productos_list,
                    'vendedores': vendedores_list,
                    'ventas_detalle': ventas_detalle,
                    'metodo_pago_breakdown': metodo_pago_list,
                }
            }

        except Exception as e:
            import traceback
            return {'success': False, 'error': str(e), 'traceback': traceback.format_exc()}

    def export_cierre_caja(self, period='today', custom_start=None, custom_end=None):
        """Genera un archivo Excel con el cierre de caja del período dado"""
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        result = self.get_profit_analysis(period, custom_start, custom_end)
        if not result['success']:
            return result

        data = result['data']
        wb = openpyxl.Workbook()

        BLUE   = PatternFill(start_color='006BA6', end_color='006BA6', fill_type='solid')
        LBLUE  = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
        WHITE_BOLD = Font(color='FFFFFF', bold=True)
        CENTER = Alignment(horizontal='center', vertical='center')
        RIGHT  = Alignment(horizontal='right')

        def header_cell(ws, row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.font = WHITE_BOLD
            c.fill = BLUE
            c.alignment = CENTER
            return c

        def auto_width(ws):
            for col in ws.columns:
                width = max((len(str(cell.value or '')) for cell in col), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 45)

        ws1 = wb.active
        ws1.title = 'Resumen'

        ws1.merge_cells('A1:C1')
        t = ws1['A1']
        t.value = f"CIERRE DE CAJA — {data['periodo']}"
        t.font = Font(color='FFFFFF', bold=True, size=13)
        t.fill = BLUE
        t.alignment = CENTER
        ws1.row_dimensions[1].height = 28

        ws1.merge_cells('A2:C2')
        g = ws1['A2']
        g.value = f"Generado: {datetime.now(BUSINESS_TZ).strftime('%d/%m/%Y %H:%M:%S')}"
        g.font = Font(italic=True, color='666666', size=9)
        g.alignment = CENTER

        stats = [
            ('Ingresos Totales',  f"${data['total_ingresos']:.2f}"),
            ('Costos Totales',    f"${data['total_costos']:.2f}"),
            ('Utilidad Neta',     f"${data['utilidad_neta']:.2f}"),
            ('Margen',            f"{data['margen_total']:.2f}%"),
            ('Total de Ventas',   data['total_ventas']),
            ('Unidades Vendidas', data['total_unidades']),
            ('Ticket Promedio',   f"${data['ticket_promedio']:.2f}"),
        ]
        for i, (label, value) in enumerate(stats):
            r = i + 4
            lc = ws1.cell(row=r, column=1, value=label)
            lc.font = Font(bold=True)
            lc.fill = LBLUE
            vc = ws1.cell(row=r, column=2, value=value)
            vc.alignment = RIGHT

        if data['vendedores']:
            vstart = len(stats) + 6
            for col, h in enumerate(['Vendedor', 'Ventas', 'Ingresos', 'Utilidad'], 1):
                header_cell(ws1, vstart, col, h)
            for i, v in enumerate(data['vendedores']):
                r = vstart + 1 + i
                ws1.cell(row=r, column=1, value=v['vendedor'])
                ws1.cell(row=r, column=2, value=v['ventas'])
                ws1.cell(row=r, column=3, value=round(v['ingresos'], 2))
                ws1.cell(row=r, column=4, value=round(v['utilidad'], 2))

        auto_width(ws1)

        ws2 = wb.create_sheet('Detalle de Ventas')
        cols2 = ['Fecha', 'Hora', 'Producto', 'Cantidad', 'Precio Venta', 'Costo Unitario', 'Ingreso', 'Costo', 'Utilidad', 'Vendedor']
        for col, h in enumerate(cols2, 1):
            header_cell(ws2, 1, col, h)

        for i, s in enumerate(data['ventas_detalle']):
            r = 2 + i
            for col, val in enumerate([
                s['fecha'], s['hora'], s['producto'], s['cantidad'],
                s['precio_venta'], s['costo_unitario'],
                s['ingreso'], s['costo'], s['utilidad'], s['vendedor']
            ], 1):
                ws2.cell(row=r, column=col, value=val)

        if data['ventas_detalle']:
            tr = 2 + len(data['ventas_detalle'])
            tc = ws2.cell(row=tr, column=3, value='TOTAL')
            tc.font = Font(bold=True)
            for col, val in [(7, data['total_ingresos']), (8, data['total_costos']), (9, data['utilidad_neta'])]:
                c = ws2.cell(row=tr, column=col, value=round(val, 2))
                c.font = Font(bold=True)

        auto_width(ws2)

        ws3 = wb.create_sheet('Top Productos')
        cols3 = ['Producto', 'Código', 'Cantidad', 'Ingresos', 'Costos', 'Utilidad']
        for col, h in enumerate(cols3, 1):
            header_cell(ws3, 1, col, h)

        for i, p in enumerate(data['productos_vendidos']):
            r = 2 + i
            for col, val in enumerate([
                p['producto'], p['codigo'], p['cantidad'],
                round(p['ingresos'], 2), round(p['costos'], 2), round(p['utilidad'], 2)
            ], 1):
                ws3.cell(row=r, column=col, value=val)

        auto_width(ws3)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return {'success': True, 'workbook': output, 'periodo': data['periodo']}
